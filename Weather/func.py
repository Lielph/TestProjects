import requests
import datetime
from openpyxl import Workbook, load_workbook
from os import path
from config import *


def convert_timezone(time, delta):
    tz = datetime.timezone(datetime.timedelta(seconds=delta))
    return datetime.datetime.fromtimestamp(time, tz=tz).strftime("%H:%M:%S")


def get_weather(city_name: str):
    params = {
        "appid": API_KEY,
        "units": UNITS,
        "lang": LANG,
        "q": city_name
    }

    try:
        r = requests.get(API_URL, params=params)
        return r.json()
    except:
        return {"cod": 0, "message": "Не удалось получить данные"}


def print_weather(data):
    if data["cod"] != 200:
        print(data["message"])
        return {}
    else:
        print("*" * 75)
        print(f"""
        Местоположение: {data["name"]},{data["sys"]["country"]}
        Сводка: {data["weather"][0]["description"]}
        Температура: {data["main"]["temp"]} °C
        Мин.температура: {data["main"]["temp_min"]} °C
        Макс.температура: {data["main"]["temp_max"]} °C
        Ощущается как: {data["main"]["feels_like"]} °C
        Давление: {data["main"]["pressure"]} кПа
        Влажность: {data["main"]["humidity"]} %
        Восход: {convert_timezone(data["sys"]["sunrise"], data["timezone"])}
        Закат: {convert_timezone(data["sys"]["sunset"], data["timezone"])}
""")
        print("*" * 75)


def save_excel(data):
    if path.exists(FILE_EXCEL):
        wb = load_workbook(filename=FILE_EXCEL)
        ws = wb.active
        ws.append([datetime.datetime.now(), f"{data["name"]},{data["sys"]["country"]}", data["main"]["temp"]])
        wb.save(FILE_EXCEL)

    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Статистика запросов"
        ws.append(["Дата запроса", "Город", "Температура"])
        ws.append([datetime.datetime.now(), f"{data["name"]},{data["sys"]["country"]}", data["main"]["temp"]])
        wb.save(FILE_EXCEL)
